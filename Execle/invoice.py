# coding=utf-8

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import shutil
import time
import os

keys = ["PurchaserName", "InvoiceDate",	"InvoiceNum", "InvoiceCode", "SellerName", "CommodityName",	"CommodityType", "CommodityUnit", "CommodityNum", "CommodityPrice","CommodityAmount","CommodityTaxRate","CommodityTax","SellerAddress","Remarks"]
titls = ["收票公司", "发票日期", "发票号码", "发票代码", "单位名称", "货物名称", "规格型号", "单位", "数量", "单价", "金额", "税率", "税额", "销售方地址及电话", "备注"]
row_len = [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]
auto_write_row  = 3;

class invoice(object):
	def __init__(self):
		basedir = os.path.join(os.environ['USERPROFILE'],"Documents")
		dir = os.path.join(basedir,"猴子读图")
		if not os.path.exists(dir):
			os.makedirs(dir) 
			
		self.filename  = "发票统计" + time.strftime('-%Y-%m-%d_%H-%M-%S',time.localtime(time.time())) + ".xlsx"
		if (os.path.exists(dir)):
			self.filename = os.path.join(dir, self.filename)

		self.wb = Workbook()
		self.ws = self.wb.active
		for i, key in enumerate(keys, 1):
			self.write(row = 1, column = i, value = key)
		
		for i, tile in enumerate(titls, 1):
			self.write(row = 2, column = i, value = tile)
		
		self.save()
	
	def save(self):
		self.set_row_width()
		self.wb.save(self.filename)
	
	def write(self, row=1, column = 1, value = ""):
		global auto_write_row
		len_tmp = len(value.encode("utf-8"))
		if (len_tmp > row_len[column - 1]):
			row_len[column - 1] = len_tmp + 2
		
		if ((row > 2) and (row >= auto_write_row)):
			auto_write_row = row + 1		
		
		d = self.ws.cell(row = row, column = column, value = value)
	
	def fill_cell(self):
		for i, key in enumerate(keys, 1):
			for j in range(4, auto_write_row):
				d = self.ws.cell(row = j, column = i)
				if d.value is None:
					d.value = self.ws.cell(row = j - 1, column = i).value

	def set_row_width(self):
		for i, row_width in enumerate(row_len, 1):
			self.ws.column_dimensions[get_column_letter(i)].width = row_width
	
	def get_auto_write_row(self):
		return auto_write_row

	def get_keys(self):
		return keys
	
	def get_file_name(self):
		return os.path.realpath(self.filename)
		
	def fix_dir(self, str="."):
		shutil.move(self.filename, str)
		self.filename = os.path.join(str,self.filename)

if __name__ == '__main__':
    ws = invoice()